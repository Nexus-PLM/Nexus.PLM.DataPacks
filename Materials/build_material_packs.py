"""Turn PLM_Materials_Library_Starter.xlsx into one Nexus data pack per material family.

Cleaning rules, applied per family:

  * Columns carrying no information are dropped: the source system's Object Id, the
    nine orthotropic columns (empty in every family, including the two where they
    would matter), Revision/Status/Modified (regenerated on import), and any column
    that is entirely blank for that family.
  * A property that is CONSTANT across all 50 materials of a family and genuinely
    varies material-to-material (hardness above all) is dropped rather than shipped:
    one hardness for fifty different steels is misinformation, not data.
  * Physically meaningless values are blanked, never kept as zero. Glass and ceramics
    are brittle: they have no yield strength, and a stored 0 reads as "zero strength".
  * Values are rounded to engineering significant figures, killing the 15-decimal
    jitter of the generated source.
  * Categorical values are mapped onto the MaterialBase LOVs (Metal/Polymer/Composite/
    Ceramic, Isotropic/Orthotropic) instead of the sheet's family names.

Provenance is preserved: every pack records that these are STARTER/TYPICAL reference
values, not certified design allowables.
"""
import csv, io, json, sys, zipfile, datetime
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

SCRATCH = Path(r"C:\Users\ma_je\AppData\Local\Temp\claude\C--Github\999fb03f-2823-456e-92ef-3993ea860581\scratchpad")
REPO = Path(r"C:\Github\Nexus.PLM.DataPacks\Materials")
XLSX = REPO / "PLM_Materials_Library_Starter.xlsx"

# sheet -> (folder Marc created, PLM type name, display name, materialCategory LOV, materialModel LOV)
FAMILIES = {
    'Steel':           ('Steel',           'n5Steel',           'Steel',            'Metal',     'Isotropic'),
    'Stainless Steel': ('Stainless Steel', 'n5StainlessSteel',  'Stainless Steel',  'Metal',     'Isotropic'),
    'Aluminium':       ('Aluminium',       'n5Aluminium',       'Aluminium',        'Metal',     'Isotropic'),
    'Titanium':        ('Titanium',        'n5Titanium',        'Titanium',         'Metal',     'Isotropic'),
    'Copper Alloys':   ('Copper Alloy',    'n5CopperAlloy',     'Copper Alloy',     'Metal',     'Isotropic'),
    'Plastics':        ('Plastic',         'n5Plastic',         'Plastic',          'Polymer',   'Isotropic'),
    'Wood':            ('Wood',            'n5Wood',            'Wood',             'Composite', 'Orthotropic'),
    'Composites':      ('Composites',      'n5Composite',       'Composite',        'Composite', 'Orthotropic'),
    'Ceramics':        ('Ceramics',        'n5Ceramic',         'Ceramic',          'Ceramic',   'Isotropic'),
    'Glass':           ('Glass',           'n5Glass',           'Glass',            'Ceramic',   'Isotropic'),
}

# Spreadsheet header -> MaterialBase attribute name.
ATTR = {
    'Material Name': 'materialName', 'Material ID / Code': 'materialCode',
    'Material Category': 'materialCategory', 'Specification / Standard': 'specificationStandard',
    'Condition / Temper': 'conditionTemper', 'Material Model': 'materialModel',
    'Reference Temperature': 'referenceTemperature', "Poisson's Ratio \u03bd": 'poissonsRatio',
    "Young's Modulus E": 'youngsModulus', 'Shear Modulus G': 'shearModulus',
    'Bulk Modulus K': 'bulkModulus', 'Density \u03c1': 'density',
    'Yield Strength \u03c3_y': 'yieldStrength',
    'Ultimate Tensile Strength \u03c3_u': 'ultimateTensileStrength',
    'Compressive Strength \u03c3_c': 'compressiveStrength',
    'Shear Strength \u03c4': 'shearStrength', 'Endurance Limit \u03c3_e': 'enduranceLimit',
    'Elongation at Break': 'elongationAtBreak', 'Hardness': 'hardness',
    'Fracture Toughness K_IC': 'fractureToughness',
    'Thermal Expansion Coefficient \u03b1': 'thermalExpansionCoeff',
    'Thermal Conductivity k': 'thermalConductivity',
    'Specific Heat Capacity c_p': 'specificHeatCapacity',
    'Melting / Softening Point T_m': 'meltingSofteningPoint',
    'Max Service Temperature': 'maxServiceTemperature',
    'Unit System': 'unitSystem', 'Data Source': 'dataSource',
}

# Never shipped: source-system identity, regenerated-on-import state, and the nine
# orthotropic columns that are empty in every family.
DROP_ALWAYS = {
    'Object Id', 'Revision', 'Status', 'Modified', 'Type',
    "Young's Modulus E1", "Young's Modulus E2", "Young's Modulus E3",
    'Shear Modulus G12', 'Shear Modulus G13', 'Shear Modulus G23',
    "Poisson's Ratio \u03bd12", "Poisson's Ratio \u03bd13", "Poisson's Ratio \u03bd23",
    # Constant per sheet and identical everywhere; folded into the pack manifest.
    'Data Status', 'Verification Notes',
}

# Constant-per-family is acceptable for these (they really are family typicals);
# for anything else, a single value shared by 50 materials is dropped as misleading.
CONSTANT_OK = {
    'Material Category', 'Material Model', 'Unit System', 'Reference Temperature',
    "Poisson's Ratio \u03bd", 'Data Source', 'Condition / Temper',
    'Specification / Standard', 'Melting / Softening Point T_m',
    'Max Service Temperature', 'Thermal Expansion Coefficient \u03b1',
    'Specific Heat Capacity c_p',
}

# Brittle families: no yield point, so a stored 0 would read as zero strength.
BRITTLE = {'Ceramics', 'Glass'}

BASE_COLUMNS = ['Part Number', 'Description', 'Type', 'Revision', 'Status', 'Source Object Id']


def sig(value, digits=4):
    """Round to `digits` significant figures, returning an int when it is whole."""
    if value is None or value == '':
        return ''
    if not isinstance(value, (int, float)):
        return str(value)
    if value == 0:
        return '0'
    from math import floor, log10
    exp = floor(log10(abs(value)))
    rounded = round(value, -(exp - digits + 1))
    if abs(rounded - round(rounded)) < 1e-9 and abs(rounded) >= 1:
        return str(int(round(rounded)))
    # Trim float noise without going scientific.
    return f'{rounded:.10f}'.rstrip('0').rstrip('.')


def to_csv(rows):
    """Same shape the app's toCsv produces: BOM, CRLF, RFC4180 quoting."""
    out = io.StringIO()
    w = csv.writer(out, lineterminator='\r\n')
    w.writerows(rows)
    return '\ufeff' + out.getvalue()


def load_type(name):
    return json.loads((SCRATCH / f'{name}.json').read_text(encoding='utf-8-sig'))


def family_type_json(template, type_name, display):
    """A concrete family type derived from MaterialBase, modelled on n5Aluminium."""
    t = json.loads(json.dumps(template))
    t['typeName'] = type_name
    t['displayName'] = display
    t['description'] = f'{display} materials.'
    t['baseTypeName'] = 'MaterialBase'
    t['abstract'] = False
    # On-disk type files store relationship as an int enum (0 = Derive), but
    # POST /api/types expects the STRING. Shipping the numeric form made the Engine
    # reject the first real import with "could not be converted to System.String".
    t['relationship'] = 'Derive'
    t['attributes'] = []          # every property is inherited from MaterialBase
    t.pop('source', None)
    return t


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    core = load_type('PLMCoreAbstractModel')
    material_base = load_type('MaterialBase')
    for t in (core, material_base):
        t['relationship'] = 'Extend' if t.get('relationship') == 1 else 'Derive'
    template = load_type('n5Aluminium')
    now = datetime.datetime.now(datetime.UTC).isoformat().replace('+00:00', 'Z')

    summary = []
    for sheet, (folder, type_name, display, category, model) in FAMILIES.items():
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

        # Decide which spreadsheet columns survive for THIS family.
        kept, dropped = [], {}
        for h in headers:
            if h in DROP_ALWAYS:
                dropped[h] = 'not shipped'
                continue
            if h in ('Part Number', 'Description'):
                continue                      # base columns, handled separately
            if h not in ATTR:
                dropped[h] = 'no MaterialBase attribute'
                continue
            values = [r[h] for r in rows]
            non_empty = [v for v in values if v not in (None, '')]
            if not non_empty:
                dropped[h] = 'empty for this family'
                continue
            if sheet in BRITTLE and h == 'Yield Strength \u03c3_y':
                dropped[h] = 'brittle: no yield point'
                continue
            if len(set(str(v) for v in values)) == 1 and h not in CONSTANT_OK:
                dropped[h] = f'one value ({sig(non_empty[0])}) shared by all {len(rows)} materials'
                continue
            kept.append(h)

        attr_names = sorted(ATTR[h] for h in kept)
        header_row = BASE_COLUMNS + [f'attr:{n}' for n in attr_names]
        by_attr = {ATTR[h]: h for h in kept}

        data_rows = []
        for r in rows:
            name = str(r['Material Name']).strip()
            values = []
            for n in attr_names:
                h = by_attr[n]
                v = r[h]
                if n == 'materialCategory':
                    v = category           # sheet held the family name, not the LOV value
                elif n == 'materialModel':
                    v = model              # 'Orthotropic / Approx.' is not an LOV value
                values.append(sig(v) if isinstance(v, (int, float)) else ('' if v is None else str(v)))
            data_rows.append([
                '',                        # Part Number: the target site mints its own
                str(r['Description']).strip(),
                type_name, '', '', '',     # Revision/Status/Source Object Id: target's own
                *values,
            ])
            _ = name

        types = [core, material_base, family_type_json(template, type_name, display)]
        manifest = {
            'formatVersion': 1,
            'name': f'{display} Materials',
            'version': '1.0',
            'author': 'Nexus PLM',
            'description': (
                f'{len(data_rows)} {display.lower()} materials with mechanical and thermal '
                'properties. STARTER / TYPICAL reference values, not certified design '
                'allowables or material certifications - verify grade, condition, temperature, '
                'test method and governing specification before design use. '
                'Values rounded to 4 significant figures. '
                'Properties that were a single value shared by every material in the family '
                'were removed rather than shipped as if they varied.'
            ),
            'sourcePrefix': 'n5',
            'exportedAt': now,
            'types': [t['typeName'] for t in types],
            'rowCounts': {type_name: len(data_rows)},
        }

        out_dir = REPO / folder
        out_dir.mkdir(parents=True, exist_ok=True)
        pack_path = out_dir / f'{folder.lower().replace(" ", "-")}-materials-1.0.nexuspack'
        csv_text = to_csv([header_row] + data_rows)

        with zipfile.ZipFile(pack_path, 'w', zipfile.ZIP_DEFLATED) as z:
            for t in types:
                z.writestr(f'types/{t["typeName"]}.json', json.dumps(t, indent=2))
            z.writestr(f'data/{type_name}.csv', csv_text)
            z.writestr('manifest.json', json.dumps(manifest, indent=2))

        # The cleaned CSV beside the pack, so the repo is reviewable without unzipping.
        (out_dir / f'{type_name}.csv').write_text(csv_text, encoding='utf-8', newline='')

        summary.append((sheet, folder, type_name, len(data_rows), len(attr_names),
                        pack_path.stat().st_size, dropped))
        print(f'{display:18} -> {pack_path.name:38} {len(data_rows):>3} rows  '
              f'{len(attr_names):>2} attrs  {pack_path.stat().st_size:>6} bytes')

    print('\n=== dropped columns per family ===')
    for sheet, folder, type_name, n, a, size, dropped in summary:
        print(f'\n{sheet}:')
        for h, why in sorted(dropped.items()):
            if why != 'not shipped':
                print(f'   - {h}: {why}')
    json.dump([{'sheet': s, 'folder': f, 'type': t, 'rows': n, 'attrs': a,
                'dropped': d} for s, f, t, n, a, _, d in summary],
              open(SCRATCH / 'pack_summary.json', 'w', encoding='utf-8'), indent=1)


if __name__ == '__main__':
    main()
